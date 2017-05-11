#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Time    : 2017-03-21 09:40
# Author  : MrFiona
# File    : insert_excel.py
# Software: PyCharm Community Edition

import os
import re
import sys
import copy
import time
import extract_data
import openpyxl
import xlsxwriter
reload(sys)
sys.setdefaultencoding('utf-8')
from cache_mechanism import DiskCache
from public_use_function import verify_validity_url, hidden_data_by_column
from setting_gloab_variable import get_url_list_by_keyword

class InsertDataIntoExcel(object):
    def __init__(self, silver_url_list, verificate_flag=False, purl_bak_string='Purley-FPGA', cache=None):
        self.cache = cache
        self.date_string_list = []
        #验证标志,默认不开启,即预测
        self.verificate_flag = verificate_flag
        #url列表初始化
        self.Silver_url_list = silver_url_list

        excel_dir = os.getcwd() + os.sep + 'excel_dir'
        if not os.path.exists(excel_dir):
            os.makedirs(excel_dir)

        self.workbook = xlsxwriter.Workbook(excel_dir + os.sep + purl_bak_string + '_report_result.xlsx', options={'strings_to_urls': False})

        if purl_bak_string == 'Purley-FPGA':
            self.rb = openpyxl.load_workbook('ITF_Skylake_FPGA_BKC_TestCase_WW19_v1.7.0_Template.xlsx', data_only=False)
        else:
            self.rb = openpyxl.load_workbook('ITF_Skylake_DE_BKC_TestCase_WW19_v1.7.0_Template.xlsx', data_only=False)

        # 在excel工作簿中增加工作表单
        self.worksheet_change = self.workbook.add_worksheet('Change History')
        self.worksheet_newsi = self.workbook.add_worksheet('NewSi')
        self.worksheet_existing = self.workbook.add_worksheet('ExistingSi')
        self.worksheet_closesi = self.workbook.add_worksheet('ClosedSi')
        self.worksheet_rework = self.workbook.add_worksheet('Rework')
        self.worksheet_rework_info = self.workbook.add_worksheet('ReworkInfo')
        self.worksheet_hw = self.workbook.add_worksheet('HW')
        self.worksheet_hw_info = self.workbook.add_worksheet('HWInfo')
        self.worksheet_sw_orignal = self.workbook.add_worksheet('SW_Original')
        self.worksheet_sw = self.workbook.add_worksheet('SW')
        self.worksheet_sw_info = self.workbook.add_worksheet('SWInfo')
        self.worksheet_ifwi_orignal = self.workbook.add_worksheet('IFWI_Original')
        self.worksheet_ifwi = self.workbook.add_worksheet('IFWI')
        self.worksheet_ifwi_info = self.workbook.add_worksheet('IFWIInfo')
        self.worksheet_platform = self.workbook.add_worksheet('ValidationResult')
        self.worksheet_mapping = self.workbook.add_worksheet('Mapping')
        self.worksheet_caseresult = self.workbook.add_worksheet('CaseResult')
        self.worksheet_save_miss = self.workbook.add_worksheet('Save-Miss')
        self.worksheet_test_time = self.workbook.add_worksheet('TestTime')

        # 为excel对象设定显示格式
        self.bold = self.workbook.add_format({'bold': 1, 'align': 'center', 'font_size': 12})
        self.bold_merge = self.workbook.add_format({'bold': 2, 'align': 'justify', 'valign': 'vcenter', 'bg_color': '#C6EFCE'})
        self.mapping_flag_style = self.workbook.add_format({'bg_color': '#EAC100'})
        # 为当前workbook添加一个样式名为titleformat
        self.titleformat = self.workbook.add_format()
        self.titleformat_header = self.workbook.add_format()
        self.titleformat_header.set_bold()  # 设置粗体字
        self.titleformat.set_font_size(12)  # 设置字体大小为10
        self.titleformat_header.set_font_size(12)  # 设置字体大小为10
        self.titleformat.set_font_name('Microsoft yahei')  # 设置字体样式为雅黑
        self.titleformat.set_align('center')  # 设置水平居中对齐
        self.titleformat_header.set_align('center')  # 设置水平居中对齐
        self.titleformat.set_align('vcenter')  # 设置垂直居中对齐
        self.titleformat_header.set_align('vcenter')  # 设置垂直居中对齐
        self.format1 = self.workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
        self.mapping_format = self.workbook.add_format({'bg_color': '#EAC100'})
        self.mapping_blue_format = self.workbook.add_format({'bg_color': '#ECF5FF'})
        self.url_format = self.workbook.add_format({'font_color': 'blue', 'underline': 1, 'align':'center', 'valign':'vcenter'})

        self.titleformat_merage = self.workbook.add_format()
        self.titleformat_merage.set_font_size(12)  # 设置字体大小为10
        self.titleformat_merage.set_align('center')  # 设置水平居中对齐
        self.titleformat_merage.set_align('vcenter')  # 设置垂直居中对齐
        # Add a format for the header cells.
        self.header_format = self.workbook.add_format({'border': 1, 'bg_color': '#C6EFCE', 'bold': True})
        #设置小数点格式部分
        self.reserve_two_decimal_format = self.workbook.add_format()
        self.reserve_zero_decimal_format = self.workbook.add_format()
        self.reserve_int_two_decimal_format = self.workbook.add_format()
        self.reserve_int_integer_format = self.workbook.add_format()
        self.reserve_int_integer_format.set_num_format( 0x01 )       #0
        self.reserve_int_two_decimal_format.set_num_format( 0x02 )   #0.00
        self.reserve_zero_decimal_format.set_num_format( 0x09 )      #0%
        self.reserve_two_decimal_format.set_num_format( 0x0a )       #0.00%

    def get_url_list(self):
        return self.Silver_url_list

    def canculate_head_num(self, multiple, num, add_num=0):
        url_length = len(self.Silver_url_list)
        # return multiple*100 - multiple*(num + 1) + add_num
        return multiple*(50 - url_length + num) + add_num

    def get_formula_data(self, data_type, wb_sheet_name):
        rb_sheet = self.rb.get_sheet_by_name(data_type)
        row, col = 1, 1
        for rs in rb_sheet.rows:
            col = 1
            for cell in rs:
                data = rb_sheet.cell(row=row, column=col).value
                if 'CaseResult' == data_type:
                    if row == 6 or (row in (2, 3) and ((col - 11) % 41) == 38) or (row == 2 and ((col - 11) % 41) == 40):
                        wb_sheet_name.write(row - 1, col - 1, data, self.reserve_zero_decimal_format)
                    elif (row in (4, 5) and ((col - 11) % 41) in (6, 7, 8, 9, 38)):
                        wb_sheet_name.write(row - 1, col - 1, data, self.reserve_int_integer_format)
                    else:
                        wb_sheet_name.write(row - 1, col - 1, data)

                elif 'Save-Miss' == data_type:
                    if 4 <= row <= 8:
                        if row != 7 and row != 6 and col == 2:
                            wb_sheet_name.write(row - 1, col - 1, data, self.reserve_two_decimal_format)
                        elif row in (6, 7) and col == 2:
                            wb_sheet_name.write(row - 1, col - 1, data, self.reserve_int_two_decimal_format)
                        elif row in (6, 7) and col != 2:
                            wb_sheet_name.write(row - 1, col - 1, data, self.reserve_int_integer_format)
                        else:
                            wb_sheet_name.write(row - 1, col - 1, data, self.reserve_zero_decimal_format)
                    else:
                        wb_sheet_name.write(row - 1, col - 1, data)

                else:
                    wb_sheet_name.write(row - 1, col - 1, data)
                col += 1
            row += 1

    def insert_change_history_data(self, purl_bak_string, k):
        self.get_formula_data('Change History', self.worksheet_change)

    def insert_Newsi_data(self, purl_bak_string, k):
        print '开始获取 New Sightings 数据........\n'
        insert_data_list_1 = ['Sighting in Test Result', 'Test Case in Test Result', 'Corresponding Test Case',
                              'Purley-FPGA No corresponding test case in test result?']

        #隐藏部分数据
        hidden_data_by_column(self.worksheet_newsi, self.Silver_url_list, 13)
        # 获取公式并插入指定位置
        self.get_formula_data(u'NewSi', self.worksheet_newsi)
        for j in range(len(self.Silver_url_list)):
            #验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
            obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
            if j == 0:
                Silver_BkC_string, date_string, effective_url_list, header_list, cell_data_list = obj.get_new_sightings_data('New Sightings', self.verificate_flag)
            else:
                Silver_BkC_string, date_string, effective_url_list, header_list, cell_data_list = obj.get_new_sightings_data('New Sightings', True)

            self.date_string_list.append(date_string)

            self.worksheet_newsi.conditional_format(4, self.canculate_head_num(13, j, 2), 250, self.canculate_head_num(13, j, 3),
                                                         {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
            self.worksheet_newsi.write(2, self.canculate_head_num(13, j, 1), Silver_BkC_string, self.format1)
            self.worksheet_newsi.write(2, self.canculate_head_num(13, j), date_string, self.format1)
            self.worksheet_newsi.write(2, self.canculate_head_num(13, j, 12), 'Comments')
            self.worksheet_newsi.write_row(2, self.canculate_head_num(13, j, 4), header_list, self.header_format)
            self.worksheet_newsi.write_row(3, self.canculate_head_num(13, j), insert_data_list_1)

            if not effective_url_list and not header_list and not cell_data_list:
                continue

            num_url_list = []
            for ele in effective_url_list:
                num_url_list.append(len(ele))

            for i in range(len(cell_data_list)):
                #插入非url部分数据
                self.worksheet_newsi.write_row(4 + i, self.canculate_head_num(13, j, 5), cell_data_list[i][1:-1])
                #插入url数据部分
                self.worksheet_newsi.write_url(4 + i, self.canculate_head_num(13, j, 4), effective_url_list[i][0], self.url_format, cell_data_list[i][0])
                if num_url_list[i] > 1:
                    self.worksheet_newsi.write_url(4 + i, self.canculate_head_num(13, j, 11), effective_url_list[i][1], self.url_format, cell_data_list[i][-1])
                elif num_url_list[i] == 1:
                    self.worksheet_newsi.write(4 + i, self.canculate_head_num(13, j, 11), cell_data_list[i][-1])

    def insert_ExistingSi_data(self, purl_bak_string, k):
        prompt_statement_list = ['Sighting in Test Result', 'Test Case in Test Result', 'Corresponding Test Case'
                                'Purley-FPGA No corresponding test case in test result?']
        # Add the standard url link format.
        fail_url_list = []

        print '开始获取  Existing Sightings 数据........\n'
        hidden_data_by_column(self.worksheet_existing, self.Silver_url_list, 13)
        # 获取公式并插入指定位置
        self.get_formula_data('ExistingSi', self.worksheet_existing)
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
            obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
            if j == 0:
                Silver_BkC_string, url_list, header_list, cell_data_list, date_string = obj.get_existing_sighting_data('Existing Sightings', self.verificate_flag)
            else:
                Silver_BkC_string, url_list, header_list, cell_data_list, date_string = obj.get_existing_sighting_data('Existing Sightings', True)

            #增加一列comments
            header_list.append('comments')
            try:
                #标记True为红色
                self.worksheet_existing.conditional_format(4, self.canculate_head_num(13, j, 2), 250, self.canculate_head_num(13, j, 3),
                                                     {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_existing.write(2, self.canculate_head_num(13, j, 1), Silver_BkC_string, self.header_format)
                self.worksheet_existing.write(2, self.canculate_head_num(13, j), date_string, self.format1)
                # 写入工作表的第二行,提示语
                self.worksheet_existing.write_row(3, self.canculate_head_num(13, j), prompt_statement_list)
                # 写入表头行
                self.worksheet_existing.write_row(2, self.canculate_head_num(13, j, 4), header_list, self.titleformat_header)
                if not url_list and not header_list and not cell_data_list:
                    continue

                # 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[7:])
                    line_num_list.append(num)
                # print line_num_list

                nu = 4
                # 插入数据第一列到第三列
                for line in range(len(cell_data_list)):
                    if line_num_list[line] <= 1:
                        self.worksheet_existing.write_row(nu, self.canculate_head_num(13, j, 5), cell_data_list[line][1:7])
                        if line_num_list[line] == 0:
                            self.worksheet_existing.write_url(nu, self.canculate_head_num(13, j, 4), url_list[line][0], self.url_format, str(cell_data_list[line][0]))
                            self.worksheet_existing.write(nu, self.canculate_head_num(13, j, 11), ' ')
                        elif line_num_list[line] == 1:
                            self.worksheet_existing.write_url(nu, self.canculate_head_num(13, j, 4), url_list[line][0], self.url_format, str(cell_data_list[line][0]))
                            self.worksheet_existing.write_url(nu, self.canculate_head_num(13, j, 11), url_list[line][1], self.url_format, str(cell_data_list[line][7]))
                        nu += 1

                    elif line_num_list[line] >= 2:
                        length_merge = line_num_list[line]
                        self.worksheet_existing.write_url(nu, self.canculate_head_num(13, j, 4), url_list[line][0], self.url_format, str(cell_data_list[line][0]))
                        self.worksheet_existing.write_row(nu, self.canculate_head_num(13, j, 5), cell_data_list[line][1:7],self.titleformat_merage)
                        for i in range(4, 11):
                            self.worksheet_existing.merge_range(nu, self.canculate_head_num(13, j, i),nu + line_num_list[line] - 1, self.canculate_head_num(13, j, i), '')
                        for m in range(length_merge):
                            self.worksheet_existing.write_url(nu + m, self.canculate_head_num(13, j, 11), url_list[line][m+1], self.url_format,str(cell_data_list[line][7 + m]))
                        nu += line_num_list[line]
            except:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
        # print '失败的url列表:\t', fail_url_list
        # print '结束获取  Existing Sightings 数据........\n'

    def insert_ClosedSi_data(self, purl_bak_string, k):
        print '开始获取 New Sightings 数据........\n'
        insert_data_list_1 = ['Sighting in Test Result', 'Test Case in Test Result', 'Corresponding Test Case',
                              'Purley-FPGA No corresponding test case in test result?']
        # 获取公式并插入指定位置

        self.get_formula_data('ClosedSi', self.worksheet_closesi)
        hidden_data_by_column(self.worksheet_closesi, self.Silver_url_list, 13)
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
            obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
            if j == 0:
                Silver_BkC_string, date_string, effective_url_list, header_list, cell_data_list = obj.get_closed_sightings_data('Closed Sightings', self.verificate_flag)
            else:
                Silver_BkC_string, date_string, effective_url_list, header_list, cell_data_list = obj.get_closed_sightings_data('Closed Sightings', True)

            self.worksheet_closesi.conditional_format(4, self.canculate_head_num(13, j, 2), 250, self.canculate_head_num(13, j, 3),
                                                      {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
            self.worksheet_closesi.write(2, self.canculate_head_num(13, j, 1), Silver_BkC_string, self.format1)
            self.worksheet_closesi.write(2, self.canculate_head_num(13, j), date_string, self.format1)
            self.worksheet_closesi.write_row(3, self.canculate_head_num(13, j), insert_data_list_1)
            self.worksheet_closesi.write_row(2, self.canculate_head_num(13, j, 4), header_list, self.header_format)
            if not effective_url_list and not header_list and not cell_data_list:
                continue

            num_url_list = []
            for ele in effective_url_list:
                num_url_list.append(len(ele))

            for i in range(len(cell_data_list)):
                #插入非url部分数据
                self.worksheet_closesi.write_row(4 + i, self.canculate_head_num(13, j, 5), cell_data_list[i][1:-1])
                #插入url数据部分
                self.worksheet_closesi.write_url(4 + i, self.canculate_head_num(13, j, 4), effective_url_list[i][0], self.url_format, cell_data_list[i][0])
                if num_url_list[i] > 1:
                    self.worksheet_closesi.write_url(4 + i, self.canculate_head_num(13, j, 11), effective_url_list[i][1], self.url_format, cell_data_list[i][-1])
                elif num_url_list[i] == 1:
                    self.worksheet_closesi.write(4 + i, self.canculate_head_num(13, j, 11), cell_data_list[i][-1])

    def insert_Rework_data(self, purl_bak_string, k):
        fail_url_list = []
        print '开始获取 HW Rework 数据........\n'
        # 获取公式并插入指定位置
        self.get_formula_data('Rework', self.worksheet_rework)

        hidden_data_by_column(self.worksheet_rework, self.Silver_url_list, 3)
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            try:
                print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
                #最新的一周在验证标志未开启的情况下取Silver数据,否则取BKC数据
                if j == 0:
                    Silver_BkC_string, object_string_list, date_string = obj.get_rework_data('HW Rework', self.verificate_flag)
                else:
                    Silver_BkC_string, object_string_list, date_string = obj.get_rework_data('HW Rework', True)

                # 标记True为红色
                self.worksheet_rework.conditional_format(3, self.canculate_head_num(3, j), 300, self.canculate_head_num(3, j),
                                                           {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_rework.write(1,self.canculate_head_num(3, j, 1), date_string, self.format1)
                self.worksheet_rework.write(1,self.canculate_head_num(3, j, 2), Silver_BkC_string, self.format1)

                if not object_string_list:
                    continue
                object_string_list.insert(0, 'TRUE')
                self.worksheet_rework.write_column(2, self.canculate_head_num(3, j, 1), object_string_list)
            except:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
        # print '失败的url列表:\t', fail_url_list
        # print '结束获取 HW Rework 数据........\n'

    def insert_ReworkInfo(self, purl_bak_string, k):
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_rework_info, self.Silver_url_list, 1)
        self.get_formula_data('ReworkInfo', self.worksheet_rework_info)

    def insert_HW_data(self, purl_bak_string, k):
        fail_url_list = []
        print '开始获取 HW Configuration 数据........\n'
        red = self.workbook.add_format({'color': 'red'})

        # 获取公式并插入指定位置
        self.get_formula_data('HW', self.worksheet_hw)
        hidden_data_by_column(self.worksheet_hw, self.Silver_url_list, 18)
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
            # 获取提取的待插入excel表的数据
            obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
            if j == 0:
                Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_hw_data('HW Configuration', self.verificate_flag)
            else:
                Silver_BkC_string, date_string, row_coordinates_list, column_coordinates_list, cell_data_list = obj.get_hw_data('HW Configuration', True)
            # Set up some formats to use.
            # 合并单元格
            self.worksheet_hw.merge_range(3, self.canculate_head_num(18, j), 9, self.canculate_head_num(18, j), "", self.titleformat_merage)
            self.worksheet_hw.write_rich_string(3, self.canculate_head_num(18, j), red, date_string, self.titleformat)

            self.worksheet_hw.write(2, self.canculate_head_num(18, j), Silver_BkC_string, self.format1)
            # 标记True为红色
            self.worksheet_hw.conditional_format(2, self.canculate_head_num(18, j, 1), 10, self.canculate_head_num(13, j, 17),
                                                 {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
            #判断这周有无数据， 例如：第40周的数据为空
            if not row_coordinates_list and not column_coordinates_list and not cell_data_list:
                fail_url_list.append(self.Silver_url_list[j])
                continue
            try:
                # 写入工作表的第二行,现在只支持system 1 到HWsystem 15，截取15行数据
                self.worksheet_hw.write_row(1, self.canculate_head_num(18, j, 3), row_coordinates_list[:15])
                #按行插入数据
                for row_num in range(len(cell_data_list)):
                    #每个元素限制为10列 [:15]
                    self.worksheet_hw.write_row(3 + row_num, self.canculate_head_num(18, j, 3), cell_data_list[row_num][:15])
            except:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
        print '失败的url列表:\t', fail_url_list
        # print '结束获取 HW Configuration 数据........\n'

    def insert_HWInfo(self, purl_bak_string, k):
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_hw_info, self.Silver_url_list, 1)
        self.get_formula_data('HWInfo', self.worksheet_hw_info)

    def insert_SW_Original_data(self, purl_bak_string, k):
        fail_url_list = []
        print '开始获取 SW_Original Configuration 数据........\n'

        red = self.workbook.add_format({'color': 'red'})
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_sw_orignal, self.Silver_url_list, 9)
        self.get_formula_data('SW_Original', self.worksheet_sw_orignal)
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            try:
                print '开始插入第[ %d ]个[ %s ]对应的数据' %(j+1, self.Silver_url_list[j])
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
                if j == 0:
                    Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list, left_col_list_1, left_col_list_2 = obj.get_sw_data('SW Configuration', self.verificate_flag)
                else:
                    Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list, left_col_list_1, left_col_list_2 = obj.get_sw_data('SW Configuration', True)

                if not header_list and not cell_data_list:
                    continue

                self.worksheet_sw_orignal.write_rich_string(1, self.canculate_head_num(9, j, 1), red, date_string, self.header_format)
                self.worksheet_sw_orignal.write_rich_string(1, self.canculate_head_num(9, j, 2), red, Silver_BkC_string, self.header_format)
                # 标记True为红色
                self.worksheet_sw_orignal.conditional_format(3, self.canculate_head_num(9, j), 40, self.canculate_head_num(9, j, 1),
                                                             {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_sw_orignal.write_row(2, self.canculate_head_num(9, j, 5), header_list, self.header_format)
                #插入数据 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[header_length - 1:])
                    line_num_list.append(num)
                nu = 3
                first_insert_data = [cell_data_list[k][0] for k in range(len(cell_data_list))]
                second_insert_data = [cell_data_list[k][2] for k in range(len(cell_data_list))]
                #插入数据第一列到第三列
                for line in range(len(cell_data_list)):
                    if line_num_list[line] == 1:
                        self.worksheet_sw_orignal.write(nu, self.canculate_head_num(9, j, 5), first_insert_data[line])
                        self.worksheet_sw_orignal.write_url(nu, self.canculate_head_num(9, j, 6), url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw_orignal.write(nu, self.canculate_head_num(9, j, 7), second_insert_data[line])

                        if (len(url_list[line]) >= 2) and cell_data_list[line][3]:
                            self.worksheet_sw_orignal.write_url(nu, self.canculate_head_num(9, j, 8), url_list[line][1], self.url_format, cell_data_list[line][3])
                        else:
                            self.worksheet_sw_orignal.write(nu, self.canculate_head_num(9, j, 8), cell_data_list[line][3])
                        nu += 1

                    elif line_num_list[line] > 1:
                        length_merge = len(url_list[line][1:])
                        self.worksheet_sw_orignal.merge_range(nu, self.canculate_head_num(9, j, 5), nu + line_num_list[line] -1, self.canculate_head_num(9, j, 5), first_insert_data[line], self.titleformat_merage)
                        self.worksheet_sw_orignal.merge_range(nu, self.canculate_head_num(9, j, 6), nu + line_num_list[line] - 1, self.canculate_head_num(9, j, 6), '')
                        self.worksheet_sw_orignal.write_url(nu, self.canculate_head_num(9, j, 6), url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw_orignal.merge_range(nu, self.canculate_head_num(9, j, 7), nu + line_num_list[line] -1, self.canculate_head_num(9, j, 7), second_insert_data[line], self.titleformat_merage)

                        for i in range(nu, length_merge + nu):
                            self.worksheet_sw_orignal.write_url(i, self.canculate_head_num(9, j, 8), url_list[line][1 + i - nu], self.url_format, cell_data_list[line][3:][i - nu])
                        self.worksheet_sw_orignal.write(line_num_list[line] + nu - 1, self.canculate_head_num(9, j, 8), cell_data_list[line][3:][-1])
                        nu += line_num_list[line]

            except:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
                # print '失败的url列表:\t', fail_url_list
                # print '结束获取 SW_Original Configuration 数据........\n'

    def insert_SW_data(self, purl_bak_string, k):
        print '开始获取 SW Configuration 数据........\n'
        red = self.workbook.add_format({'color': 'red'})
        fail_url_list = []

        #获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_sw, self.Silver_url_list, 9)
        self.get_formula_data('SW', self.worksheet_sw)
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            try:
                print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
                if j == 0:
                    Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list, left_col_list_1, left_col_list_2 = obj.get_sw_data('SW Configuration', self.verificate_flag)
                else:
                    Silver_BkC_string, header_length, date_string, url_list, header_list, cell_data_list, left_col_list_1, left_col_list_2 = obj.get_sw_data('SW Configuration', True)

                if not header_list and not cell_data_list:
                    continue

                self.worksheet_sw.write_rich_string(1, self.canculate_head_num(9, j, 1), red, date_string, self.header_format)
                self.worksheet_sw.write_rich_string(1, self.canculate_head_num(9, j, 2), red, Silver_BkC_string, self.header_format)
                # Set up some formats to use.
                self.worksheet_sw.conditional_format(3, self.canculate_head_num(9, j), 40, self.canculate_head_num(9, j, 1),
                                                     {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_sw.write_row(2, self.canculate_head_num(9, j, 5), header_list, self.header_format)
                # 插入数据 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[header_length - 1:])
                    line_num_list.append(num)

                #处理并插入数据
                insert_data = zip( cell_data_list, url_list )
                insert_data.sort(key=lambda x: x[0][0].upper())

                cell_data_list, url_list = zip( *insert_data )

                first_insert_data = [cell_data_list[k][0] for k in range(len(cell_data_list))]
                second_insert_data = [cell_data_list[k][2] for k in range(len(cell_data_list))]
                # 插入数据第一列到第三列
                nu = 3
                for line in range(len(cell_data_list)):
                    if line_num_list[line] == 1:
                        self.worksheet_sw.write(nu, self.canculate_head_num(9, j, 5), first_insert_data[line])
                        self.worksheet_sw.write_url(nu, self.canculate_head_num(9, j, 6), url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw.write(nu, self.canculate_head_num(9, j, 7), second_insert_data[line])

                        if (len(url_list[line]) >= 2) and cell_data_list[line][3]:
                            self.worksheet_sw.write_url(nu, self.canculate_head_num(9, j, 8), url_list[line][1], self.url_format, cell_data_list[line][3])
                        else:
                            self.worksheet_sw.write(nu, self.canculate_head_num(9, j, 8), cell_data_list[line][3])
                        nu += 1

                    elif line_num_list[line] > 1:
                        length_merge = len(url_list[line][1:])
                        self.worksheet_sw.merge_range(nu, self.canculate_head_num(9, j, 5),
                                                      nu + line_num_list[line] - 1, self.canculate_head_num(9, j, 5), first_insert_data[line], self.titleformat_merage)
                        self.worksheet_sw.merge_range(nu, self.canculate_head_num(9, j, 6),
                                                      nu + line_num_list[line] - 1, self.canculate_head_num(9, j, 6), '')
                        self.worksheet_sw.write_url(nu, self.canculate_head_num(9, j, 6),
                                                    url_list[line][0], self.url_format, cell_data_list[line][1])
                        self.worksheet_sw.merge_range(nu, self.canculate_head_num(9, j, 7),
                                                      nu + line_num_list[line] - 1, self.canculate_head_num(9, j, 7), second_insert_data[line], self.titleformat_merage)
                        for i in range(nu, length_merge + nu):
                            self.worksheet_sw.write_url(i, self.canculate_head_num(9, j, 8),
                                                        url_list[line][1 + i - nu], self.url_format, cell_data_list[line][3:][i - nu])
                        self.worksheet_sw.write(line_num_list[line] + nu - 1, self.canculate_head_num(9, j, 8), cell_data_list[line][3:][-1])
                        nu += line_num_list[line]

            except:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
        # print '失败的url列表:\t', fail_url_list
        # print '结束获取 SW Configuration 数据........\n'

    def insert_SWInfo(self, purl_bak_string, k):
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_sw_info, self.Silver_url_list, 1)
        self.get_formula_data('SWInfo', self.worksheet_sw_info)

    def insert_IFWI_Orignal_data(self, purl_bak_string, k):
        print '开始获取 IFWI_Orignal Configuration 数据........\n'
        red = self.workbook.add_format({'color': 'red'})
        blue = self.workbook.add_format({'color': 'blue'})
        fail_url_list = []

        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_ifwi_orignal, self.Silver_url_list, 6)
        self.get_formula_data('IFWI_Original', self.worksheet_ifwi_orignal)
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            try:
                print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
                if j == 0:
                    Silver_BkC_string, date_string, header_list, cell_data_list = obj.get_ifwi_data('IFWI Configuration', self.verificate_flag)
                else:
                    Silver_BkC_string, date_string, header_list, cell_data_list = obj.get_ifwi_data('IFWI Configuration', True)
                # Set up some formats to use.
                self.worksheet_ifwi_orignal.conditional_format(4, self.canculate_head_num(6, j), 50, self.canculate_head_num(6, j, 1),
                                                             {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_ifwi_orignal.write_rich_string(1, self.canculate_head_num(6, j, 1), red, date_string, self.header_format)
                self.worksheet_ifwi_orignal.write_rich_string(1, self.canculate_head_num(6, j, 2), red, Silver_BkC_string, self.format1)

                if not header_list and not cell_data_list:
                    continue

                self.worksheet_ifwi_orignal.write_row(3, self.canculate_head_num(6, j, 3), header_list, self.header_format)
                #插入数据,需要考虑有数字的情况，前面加Nic字母
                for ele in range(len(cell_data_list)):
                    #以数字开头的元素前面加Nic
                    match_obj = re.match('\s+\d+', str(cell_data_list[ele][0]))
                    if match_obj:
                        cell_data_list[ele][0] = 'Nic' + cell_data_list[ele][0]
                    cell_data_list[ele][0] = cell_data_list[ele][0].lstrip(' ')

                for begain in range(len(cell_data_list)):
                    self.worksheet_ifwi_orignal.write_row(begain + 4,  self.canculate_head_num(6, j, 3), cell_data_list[begain])
            except ValueError:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
        # print '失败的url列表:\t', fail_url_list
        # print '结束获取 IFWI_Orignal Configuration 数据........\n'

    def insert_IFWI_data(self, purl_bak_string, k):
        print '开始获取 IFWI Configuration 数据........\n'
        red = self.workbook.add_format({'color': 'red'})
        fail_url_list = []

        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_ifwi, self.Silver_url_list, 6)
        self.get_formula_data('IFWI', self.worksheet_ifwi)
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            try:
                print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
                if j == 0:
                    Silver_BkC_string, date_string, header_list, cell_data_list = obj.get_ifwi_data('IFWI Configuration',self.verificate_flag)
                else:
                    Silver_BkC_string, date_string, header_list, cell_data_list = obj.get_ifwi_data('IFWI Configuration', True)

                if not header_list and not cell_data_list:
                    continue

                self.worksheet_ifwi.write_rich_string(1, self.canculate_head_num(6, j, 1), red, date_string, self.header_format)
                self.worksheet_ifwi.write_rich_string(1, self.canculate_head_num(6, j, 2), red, Silver_BkC_string, self.format1)
                # Set up some formats to use.
                self.worksheet_ifwi.conditional_format(4, self.canculate_head_num(6, j), 50, self.canculate_head_num(6, j, 1),
                                                       {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
                self.worksheet_ifwi.write_row(4, self.canculate_head_num(6, j, 3), header_list, self.header_format)
                #插入数据,需要考虑有数字的情况，前面加Nic字母
                for ele in range(len(cell_data_list)):
                    #以数字开头的元素前面加Nic
                    match_obj = re.match('^\d+', str(cell_data_list[ele][0]))
                    if match_obj:
                        cell_data_list[ele][0] = 'Nic' + cell_data_list[ele][0]
                    cell_data_list[ele][0] = cell_data_list[ele][0].lstrip(' ')

                cell_data_list.sort(key=lambda x: x[0].upper())
                for begain in range(len(cell_data_list)):
                    self.worksheet_ifwi.write_row(begain + 4,  self.canculate_head_num(6, j, 3), cell_data_list[begain])
            except ValueError:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
        # print '失败的url列表:\t', fail_url_list
        # print '结束获取 IFWI Configuration 数据........\n'

    def insert_IFWIInfo(self, purl_bak_string, k):
        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_ifwi_info, self.Silver_url_list, 1)
        self.get_formula_data('IFWIInfo', self.worksheet_ifwi_info)

    def insert_Platform_data(self, purl_bak_string, k):
        print '开始获取 Platform Integration Validation Result 数据........\n'
        # Set up some formats to use.
        red = self.workbook.add_format({'color': 'red'})

        # 获取公式并插入指定位置
        hidden_data_by_column(self.worksheet_platform, self.Silver_url_list, 12)
        self.get_formula_data('ValidationResult', self.worksheet_platform)
        fail_url_list = []
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            try:
                print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
                if j == 0:
                    Silver_BkC_string, date_string, url_list, header_list, cell_data_list = obj.get_platform_data('Platform Integration Validation Result', self.verificate_flag)
                else:
                    Silver_BkC_string, date_string, url_list, header_list, cell_data_list = obj.get_platform_data('Platform Integration Validation Result', True)

                if not header_list and not cell_data_list:
                    self.worksheet_platform.write_rich_string(2, self.canculate_head_num(12, j), red, date_string, self.titleformat)
                    continue

                self.worksheet_platform.write_rich_string(1, self.canculate_head_num(12, j, 1), red, Silver_BkC_string, self.format1)

                self.worksheet_platform.write_row(2, self.canculate_head_num(12, j, 1), header_list, self.header_format)
                # 最后一列可能会出现多值多行的情况，计算每行数据占有的行数
                line_num_list = []
                for ele in cell_data_list:
                    num = len(ele[10:])
                    line_num_list.append(num)
                #有可能出现多个多行的情况，自适应合并列的变化
                merge_width = len(cell_data_list) + 2
                for k in line_num_list:
                    if k > 1:
                        merge_width += k - 1

                self.worksheet_platform.merge_range(2, self.canculate_head_num(12, j), merge_width, self.canculate_head_num(12, j), '', self.titleformat_merage)
                self.worksheet_platform.write_rich_string(2, self.canculate_head_num(12, j), red, date_string, self.titleformat)
                # 标记True为红色
                self.worksheet_platform.conditional_format(3, self.canculate_head_num(12, j, 5), len(cell_data_list) + 3, self.canculate_head_num(12, j, 5),
                                                           {'type': 'cell', 'criteria': '>', 'value': 0, 'format': self.format1})
                nu = 3
                # 插入数据第一列到第三列
                for line in range(len(cell_data_list)):
                    #将数字字符转化为数字
                    temp_list = copy.deepcopy(cell_data_list[line])
                    for ele in range(len(temp_list)):
                        if temp_list[ele].isdigit():
                            temp_list[ele] = int(temp_list[ele])

                    if line_num_list[line] == 1:
                        self.worksheet_platform.write_row(nu, self.canculate_head_num(12, j, 1), temp_list)
                        if url_list[line]:
                            self.worksheet_platform.write_url(nu, self.canculate_head_num(12, j, 11), url_list[line][0], self.url_format, str(temp_list[10]))
                        nu += 1

                    elif line_num_list[line] > 1:
                        length_merge = line_num_list[line]
                        for i in range(1, 11):
                            self.worksheet_platform.merge_range(nu, self.canculate_head_num(12, j, i),nu + line_num_list[line] - 1,self.canculate_head_num(12, j, i), '')
                        self.worksheet_platform.write_row(nu, self.canculate_head_num(12, j, 1), temp_list[:10], self.titleformat_merage)
                        for m in range(length_merge):
                            self.worksheet_platform.write_url(nu+m, self.canculate_head_num(12, j, 11), url_list[line][m], self.url_format, str(temp_list[10+m]))
                        nu += line_num_list[line]

            except:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
        # print '失败的url列表:\t', fail_url_list
        # print '结束获取 Platform Integration Validation Result 数据........\n'

    def insert_Mapping(self, purl_bak_string, k):
        height = 50 - len(self.Silver_url_list) + 5
        for row in range(height - 5):
            self.worksheet_mapping.set_row(row=row + 6, options={'hidden':True})

        self.worksheet_mapping.set_column(5, height - 1, options={'hidden': True})

        # 获取公式并插入指定位置
        self.get_formula_data('Mapping', self.worksheet_mapping)
        # first_value_list = ['Considering SW Ingredient adding as change?'] + ['TRUE']*100
        # second_value_list = first_value_list
        # third_value_list = ['Considering SW Ingredient adding as change?'] + ['FALSE']*100
        # fourth_value_list = third_value_list
        # 标记True为红色
        self.worksheet_mapping.conditional_format(57, 5, 200, 56, {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})
        self.worksheet_mapping.conditional_format(57, 57, 200, 91, {'type': 'cell', 'criteria': '=', 'value': 1, 'format': self.format1})
        # self.worksheet_mapping.write_row(0, 4, first_value_list, self.mapping_flag_style)
        # self.worksheet_mapping.write_row(1, 4, second_value_list, self.mapping_flag_style)
        # self.worksheet_mapping.write_row(2, 4, third_value_list, self.mapping_flag_style)
        # self.worksheet_mapping.write_row(3, 4, fourth_value_list, self.mapping_flag_style)
        insert_date_list = []
        orignal_url_list = self.Silver_url_list
        for url in orignal_url_list:
            split_string_list = str(url).split('/')
            object_string = split_string_list[-2][-4:]
            insert_date_list.append(object_string)
        self.worksheet_mapping.write_row(56, 56 - len(self.Silver_url_list) - 1, insert_date_list + ['EXX_WW10', 'changed?\Test?'], self.header_format)
        self.worksheet_mapping.write_column(56 - len(self.Silver_url_list) - 1, 56, insert_date_list, self.header_format)

        row_header_list = []
        row_value_list = ['Power Management', 'Networking', 'USB', 'FPGA', 'Video', 'Storage', 'PCIe', 'Manageability',
                           'Virtualization', 'Memory', 'Stress & Stability', 'Security', 'Processor', 'DPDK', 'DPDK_FVL_10G',
                           'QAT', 'RAS', 'Summary']
        for head in row_value_list:
            row_header_list.append(head)
            if head == 'RAS' or head == 'Summary':
                continue
            row_header_list.append(' ')
        self.worksheet_mapping.write_row(5, 58, row_header_list, self.header_format)

    def insert_CaseResult(self, purl_bak_string, k):
        print '开始获取 CaseResult 数据........\n'
        fail_url_list = []

        self.worksheet_caseresult.set_column(firstcol=10, lastcol=41 * (50 - len(self.Silver_url_list)) + 10, options={'hidden': True})

        # 获取公式并插入指定位置
        self.get_formula_data('CaseResult', self.worksheet_caseresult)
        yellow_header_format = self.workbook.add_format({'bg_color': '#FFFF66'})
        brown_header_format = self.workbook.add_format({'bg_color': '#FF9933'})
        dark_red_header_format = self.workbook.add_format({'bg_color': '#FF2D2D'})
        green_header_format = self.workbook.add_format({'bg_color': '#79FF79'})
        blue_header_format = self.workbook.add_format({'bg_color': '#97CBFF'})
        blue = self.workbook.add_format({'color': 'blue', 'font_size': 13})
        caseresult_header_format = self.workbook.add_format({'border': 1, 'align': 'center', 'bg_color': '#97CBFF', 'bold': 1, 'font_size': 13})
        caseresult_last_header_format = self.workbook.add_format({'bold': 1, 'align': 'center', 'font_size': 13})
        start_string = ['START']*31
        yellow_promote_string = ['Reported FAILED Sightings in test result', 'Reported Sightings in test result',
                                 'Missed Key Sightings', "Missed Key Sightings which is not due to new added test case.  "
                                "If this issue can be found by other selected test case, consider it as not missed.  "
                                "If this issue failed one test case, and blocked some other test cases, then those blocked test case don't have to be selected.",
                                 'Missed Sightings in test plan comparing to test result?', 'Missed Test Case in Test Plan?',
                                 'New added Test Case comparing to Test Plan Pool?', 'Different Test Case between Test Plan and test result?',
                                 'New added Test Case comparing to entire Test Pool?']
        brown_promote_string = ['Removed Test Case in Test Result?', 'Different Test Case between Test Plan Pool and previous Test Result?',
                                'New Sighting with old test case in test result?', 'Missed New Sighting with old test case in test plan?',
                                'ExistingSighting in previous test result?', 'Fixed Sighting in release notes?',
                                'Fixed Sighting in test result?', 'Missed Fixed Sighting in test plan?',
                                'New Add Test Case comparing to previous test plan pool?', 'Mapped in Mapping table?',
                                'NOT Found in Mapping table?', 'Covered new Sighting?', 'Covered exsiting sighting?',
                                'Covered Sighting?', 'Missed Sighting in test plan', 'No run in the past x(x<=10) weeks?',
                                'No Issue found in the latest y(y<=10) tests?', 'Basic Case?', 'Efforts override',
                                'Efforts', 'Not in actual test result but is selected in Test Plan.', 'Selected(z = threshold value)?']
        last_header_list = ['Domain', 'Category', 'Case', 'New Added Test Case Comparing to Previous Test Case Pool',
                            'New Added Test Case Comparing to Entire Test Case Pool']
        for j in range(len(self.Silver_url_list)):
            # 验证url有效性
            is_validity = verify_validity_url(self.Silver_url_list[j])
            if not is_validity:
                continue
            try:
                print '开始插入第[ %d ]个[ %s ]对应的数据' % (j + 1, self.Silver_url_list[j])
                obj = extract_data.GetAnalysisData(self.Silver_url_list[j], get_info_not_save_flag=False, cache=self.cache)
                if j == 0:
                    date_string, tip_string, header_list, cell_data_list = obj.get_caseresult_data('Platform Integration Validation Result', self.verificate_flag)
                else:
                    date_string, tip_string, header_list, cell_data_list = obj.get_caseresult_data('Platform Integration Validation Result', True)

                if not header_list and not cell_data_list:
                    continue

                #按照格式写入一些数据
                effective_tip_string = tip_string.split()[0] + ' TestResult ' + tip_string.split()[1]
                effective_tip_last_string = tip_string.split()[0] + ' Plan ' + tip_string.split()[1]
                self.worksheet_caseresult.write_rich_string(6, self.canculate_head_num(41, j, 1+11), blue, effective_tip_string)
                self.worksheet_caseresult.write_rich_string(6, self.canculate_head_num(41, j, 36+1+11), blue, effective_tip_last_string)

                for i in range(len(yellow_promote_string)):
                    self.worksheet_caseresult.write_rich_string(6, self.canculate_head_num(41, j, 5+i+11), yellow_promote_string[i], yellow_header_format)
                    self.worksheet_caseresult.write_rich_string(7, self.canculate_head_num(41, j, 5+i+11), start_string[:9][i], yellow_header_format)
                for i in range(len(brown_promote_string)):
                    self.worksheet_caseresult.write_rich_string(6, self.canculate_head_num(41, j, 14+i+11), brown_promote_string[i], brown_header_format)
                self.worksheet_caseresult.write_row(7, self.canculate_head_num(41, j, 14+11), start_string[9:])

                self.worksheet_caseresult.write_row(7, self.canculate_head_num(41, j, 11), header_list, caseresult_header_format)
                self.worksheet_caseresult.write_row(7, self.canculate_head_num(41, j, 36+11), last_header_list, caseresult_last_header_format)
                # 标记True为红色
                self.worksheet_caseresult.conditional_format(8, self.canculate_head_num(41, j, 3+11), 250, self.canculate_head_num(41, j, 3+11),
                                                             {'type': 'text', 'criteria': 'containing', 'value': 'FAILED', 'format': dark_red_header_format})
                self.worksheet_caseresult.conditional_format(8, self.canculate_head_num(41, j, 3+11), 250, self.canculate_head_num(41, j, 3+11),
                                                             {'type': 'text', 'criteria': 'containing', 'value': 'PASSED', 'format': green_header_format})
                self.worksheet_caseresult.conditional_format(8, self.canculate_head_num(41, j, 13+11), 250, self.canculate_head_num(41, j, 13+11),
                                                             {'type': 'cell', 'criteria': '=', 'value': False, 'format': blue_header_format})
                self.worksheet_caseresult.conditional_format(8, self.canculate_head_num(41, j, 23+11), 250, self.canculate_head_num(41, j, 23+11),
                                                             {'type': 'cell', 'criteria': '=', 'value': 0, 'format': self.format1})
                for line in range(len(cell_data_list)):
                    self.worksheet_caseresult.write_row(8 + line, self.canculate_head_num(41, j, 11), cell_data_list[line])

                self.worksheet_caseresult.conditional_format(8, self.canculate_head_num(41, j, 9+11), 250, self.canculate_head_num(41, j, 35+11),
                                                             {'type': 'cell', 'criteria': '=', 'value': True, 'format': self.format1})

            except:
                print '爬取第[ %d ]个[ %s ]对应的数据失败' % (j + 1, self.Silver_url_list[j])
                fail_url_list.append(self.Silver_url_list[j])
        # print '失败的url列表:\t', fail_url_list
        # print '结束获取 CaseResult 数据........\n'

    def insert_save_miss_data(self, purl_bak_string, k):
        self.worksheet_save_miss.set_column(firstcol=2, lastcol=50 - len(self.Silver_url_list) + 1, options={'hidden': True})
        # 获取公式并插入指定位置
        yellow_header_format = self.workbook.add_format({'bg_color': '#FFFF66'})
        first_string_list = ["Save Test Case based on this week's Test case pool", "Miss Sightings in test plan comparing to test result (%)?",
                             "Miss Sightings in test plan comparing to test result?", "Total Sighting", "Save Effort"]
        self.get_formula_data('Save-Miss', self.worksheet_save_miss)
        self.worksheet_save_miss.write_row(2, 52 - len(self.Silver_url_list), self.date_string_list, self.format1)
        for i in range(len(first_string_list)):
            self.worksheet_save_miss.write_rich_string(3 + i, 0, first_string_list[i], yellow_header_format)
        # for i in range(1, 101):
        #     self.worksheet_save_miss.write_rich_string(0, 102 - i, str(i), yellow_header_format)
        # self.worksheet_save_miss.write(12, 0, 11.0, self.reserve_int_two_decimal_format)
        # self.worksheet_save_miss.write(13, 0, 11.0, self.reserve_int_integer_format)

    def insert_test_time_data(self, purl_bak_string, k):
        # 获取公式并插入指定位置
        self.get_formula_data('TestTime', self.worksheet_test_time)

    def return_name(self):
        return self.__class__.__dict__

    def close_workbook(self):
        self.workbook.close()


if __name__ == '__main__':
    from public_use_function import Silver_url_list
    start = time.time()
    cache = DiskCache()
    obj = InsertDataIntoExcel(verificate_flag=False, cache=cache, silver_url_list=Silver_url_list)
    # obj.insert_Mapping('Purley-FPGA', 1)
    # obj.insert_save_miss_data('Purley-FPGA', 1)
    # # getattr(obj, 'insert_ExistingSi_data')()
    func_name_list = obj.return_name().keys()
    use_func_list = [func for func in func_name_list if func.startswith('insert')]
    print use_func_list
    for func in use_func_list:
        getattr(obj, func)('Purley-FPGA', 1)
    obj.close_workbook()

    print time.time() - start